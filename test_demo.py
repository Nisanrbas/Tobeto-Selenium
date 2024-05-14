from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from constants import cons as c




class Test_DemoClass:
    def setup_method(self):
       self.driver = webdriver.Chrome()
       self.driver.maximize_window() 
       self.driver.get(c.URL)
       self.folderPath=str(date.today())
       Path(self.folderPath).mkdir(exist_ok=True)
     
       #günün tarihi al bu güne ait kloser varsa o yoksa oluştur

    def teardown_method(self):#her test sonunda 
        self.driver.quit()#her testen sonra tarayıcı kapat



    # @pytest.mark.skip() bu testi atla demek

    def getData():
        #dekaratör den çağırılan fonksiyona self verilmez
        
        excelFile=openpyxl.load_workbook("data\log.xlsx")
        selectedSheet=excelFile["Sayfa1"]
        totalsrow=selectedSheet.max_row
        data=[]
        for i in range(2,totalsrow+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleDate=(username,password)
            data.append(tupleDate)
        return data
    @pytest.mark.parametrize("username,password",getData())
     #kullanıcı olmaması hatası
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,c.loginbtn)
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,c.erormessage)
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        #magic string
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    @pytest.mark.parametrize("username,password",[("locked_out_user","secret_sauce" )]) 

    def test_invalid2_login(self,username,password):

        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,c.loginbtn)
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,c.erormessage)
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        #magic string
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."

    @pytest.mark.parametrize("username,password",[("secret_sauce",""),("Nisa","")]) 

    def test_emptyPassword_login(self,username,password):

        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,c.loginbtn)
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,c.erormessage)
        
        #magic string
        assert errorMessage.text == "Epic sadface: Password is required"

    @pytest.mark.parametrize("username,password",[("","" )])

    def test_emptyUsernamePassword_login(self,username,password):

        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,c.loginbtn)
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,c.erormessage)
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        #magic string
        assert errorMessage.text == "Epic sadface: Username is required"

    @pytest.mark.parametrize("username,password",[("standard_use","secret_sauce")])

    def test_url_login(self,username,password):

        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,c.loginbtn)
        loginBtn.click()
        sleep(3)

        link=self.driver.current_url
      
      # Giriş yapıldıktan sonra kullanıcıya gösterilen ürün sayısı "6" adet olmalıdır.
    @pytest.mark.parametrize("username,password",[("standard_user","secret_sauce")])
    def test_numberofproduct(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput=self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        passwordInput=self.driver.find_element(By.ID,"password")
        
        action=ActionChains(self.driver)
        action.send_keys_to_element(usernameInput,username)
        action.send_keys_to_element(passwordInput,password)
        action.perform()

        loginBtn=self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        
       
        listofproduct=self.driver.find_elements(By.CLASS_NAME,"inventory_item")
        assert len(listofproduct)==6

       


    #Sepete ürün eklenmelidir.
    @pytest.mark.parametrize("username,password",[("problem_user","secret_sauce")])
    def test_addproduct(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput=self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"))
        passwordInput=self.driver.find_element(By.ID,"password")
        
        action=ActionChains(self.driver)
        action.send_keys_to_element(usernameInput,username)
        action.send_keys_to_element(passwordInput,password)
        action.perform()

        loginBtn=self.driver.find_element(By.ID,"login-button")
        loginBtn.click()

        addProduct=self.driver.find_element(By.XPATH,"/html/body/div[1]/div/div/div[2]/div/div/div/div[3]/div[2]/div[2]/button")
        addProduct.click()

       
 
      



    



    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
  